import os
import uuid
import logging
from datetime import datetime
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

logger = logging.getLogger(__name__)


def generate_excel(analysis: str, transcript: str = None) -> str:
    """
    Генерация Excel файла с анализом встречи
    
    Args:
        analysis: Текст анализа от GPT
        transcript: Полный транскрипт (опционально)
    
    Returns:
        Путь к созданному файлу
    """
    try:
        wb = Workbook()
        
        # === Лист 1: Summary ===
        ws_summary = wb.active
        ws_summary.title = "Summary"
        
        # Заголовок
        ws_summary['A1'] = "Анализ встречи"
        ws_summary['A1'].font = Font(size=16, bold=True)
        ws_summary['A2'] = f"Дата: {datetime.now().strftime('%Y-%m-%d %H:%M')}"
        
        # Извлечение summary из анализа
        summary = extract_section(analysis, "Summary")
        ws_summary['A4'] = "Резюме:"
        ws_summary['A4'].font = Font(bold=True)
        ws_summary['A5'] = summary
        ws_summary['A5'].alignment = Alignment(wrap_text=True, vertical='top')
        
        # === Лист 2: Задачи ===
        ws_tasks = wb.create_sheet("Задачи")
        
        # Заголовки таблицы
        headers = ["№", "Задача", "Дедлайн", "Ответственный", "Приоритет", "Статус"]
        for col, header in enumerate(headers, 1):
            cell = ws_tasks.cell(1, col, header)
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
            cell.alignment = Alignment(horizontal='center')
        
        # Парсинг задач из анализа
        tasks = parse_tasks_from_analysis(analysis)
        
        for row, task in enumerate(tasks, 2):
            ws_tasks.cell(row, 1, row - 1)
            ws_tasks.cell(row, 2, task.get('description', ''))
            ws_tasks.cell(row, 3, task.get('deadline', 'Не указан'))
            ws_tasks.cell(row, 4, task.get('assignee', 'Не указан'))
            ws_tasks.cell(row, 5, task.get('priority', 'Средний'))
            ws_tasks.cell(row, 6, "Новая")
        
        # Автоширина колонок
        for col in range(1, 7):
            max_length = 0
            column = get_column_letter(col)
            for cell in ws_tasks[column]:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(cell.value)
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            ws_tasks.column_dimensions[column].width = adjusted_width
        
        # === Лист 3: Полный анализ ===
        ws_full = wb.create_sheet("Полный анализ")
        ws_full['A1'] = analysis
        ws_full['A1'].alignment = Alignment(wrap_text=True, vertical='top')
        ws_full.column_dimensions['A'].width = 100
        
        # === Лист 4: Транскрипт (если есть) ===
        if transcript:
            ws_transcript = wb.create_sheet("Транскрипт")
            ws_transcript['A1'] = transcript
            ws_transcript['A1'].alignment = Alignment(wrap_text=True, vertical='top')
            ws_transcript.column_dimensions['A'].width = 100
        
        # Сохранение файла
        file_id = str(uuid.uuid4())
        filename = f"analysis_{file_id}.xlsx"
        file_path = f"/tmp/{filename}"
        
        wb.save(file_path)
        logger.info(f"Excel file created: {file_path}")
        
        return file_path
    
    except Exception as e:
        logger.error(f"Excel generation failed: {e}")
        raise Exception(f"Не удалось создать Excel файл: {str(e)}")


def extract_section(text: str, section_name: str) -> str:
    """Извлечение секции из текста"""
    lines = text.split('\n')
    result = []
    in_section = False
    
    for line in lines:
        if section_name.lower() in line.lower() and '#' in line:
            in_section = True
            continue
        if in_section:
            if line.strip().startswith('##'):
                break
            if line.strip():
                result.append(line.strip())
    
    return '\n'.join(result) if result else "Не найдено"


def parse_tasks_from_analysis(analysis: str) -> list:
    """Парсинг задач из анализа"""
    tasks = []
    lines = analysis.split('\n')
    
    current_task = None
    for line in lines:
        if 'Задача:' in line:
            if current_task:
                tasks.append(current_task)
            
            current_task = {
                'description': line.split('Задача:')[1].strip(),
                'deadline': 'Не указан',
                'assignee': 'Не указан',
                'priority': 'Средний'
            }
        elif current_task:
            if 'Дедлайн:' in line:
                current_task['deadline'] = line.split('Дедлайн:')[1].strip()
            elif 'Ответственный:' in line:
                current_task['assignee'] = line.split('Ответственный:')[1].strip()
            elif 'Приоритет:' in line:
                current_task['priority'] = line.split('Приоритет:')[1].strip()
    
    if current_task:
        tasks.append(current_task)
    
    return tasks
